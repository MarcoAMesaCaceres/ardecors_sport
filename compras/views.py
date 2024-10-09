from django.shortcuts import get_object_or_404, redirect, render
from .models import Compras
from .forms import ComprasForm
from django.db.models import Q
from django.template.loader import get_template

# Exportar pdf, excel
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db import transaction
from django.core.exceptions import ValidationError
from django.db.models import Q


def lista_compras(request):
    compras = Compras.objects.all()
    return render(request, 'lista_compras.html', {'compras': compras})

def crear_compras(request):
    if request.method == 'POST':
        form = ComprasForm(request.POST)
        if form.is_valid():
            try:
                compra = form.save()
                messages.success(request, 'Compra creada exitosamente.')
                return redirect('crear_detalle_compra', compra_id=compra.id)
            except ValidationError as e:
                messages.error(request, f'Error al crear la compra: {e}')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{field}: {error}')
    else:
        form = ComprasForm()
    return render(request, 'crear_compras.html', {'form': form})

def editar_compras(request, pk):
    compra = get_object_or_404(Compras, pk=pk)
    if request.method == 'POST':
        form = ComprasForm(request.POST, instance=compra)
        if form.is_valid():
            form.save()
            return redirect('lista_compras')
    else:
        form = ComprasForm(instance=compra)
    return render(request, 'editar_compras.html', {'form': form, 'compra': compra})

def eliminar_compras(request, pk):
    compra = get_object_or_404(Compras, pk=pk)  # Cambiado de 'compras' a 'Compras'
    if request.method == 'POST':
        compra.delete()
        return redirect('lista_compras')  # Cambiado aquí
    return render(request, 'eliminar_compras.html', {'compra': compra})

def exportar_excel(request):
    compras = Compras.objects.all()
    
    # Crear libro y hoja de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Compras'

    # Agregar encabezados
    headers = ['ID', 'Fecha', 'Proveedor', 'Producto', 'Cantidad', 'Precio Unitario', 'Total']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Agregar filas de compras
    for row_num, compra in enumerate(compras, 2):
        ws[f'A{row_num}'] = compra.id
        ws[f'B{row_num}'] = compra.fecha.strftime('%d/%m/%Y')
        ws[f'C{row_num}'] = compra.proveedor.nombre if compra.proveedor else 'No especificado'
        ws[f'D{row_num}'] = compra.producto
        ws[f'E{row_num}'] = 'No aplica'  # Asume que no se tiene cantidad en el modelo
        ws[f'F{row_num}'] = 'No aplica'  # Asume que no se tiene precio_unitario en el modelo
        ws[f'G{row_num}'] = compra.total

    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=compras.xlsx'

