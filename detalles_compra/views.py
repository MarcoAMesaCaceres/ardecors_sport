from django.shortcuts import render, get_object_or_404, redirect
from .forms import DetalleCompraForm
from .models import DetalleCompra
from compras.models import Compras  # Añade esta importación

# Exportar pdf, excel
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle




def lista_detalles_compra(request, compra_id=None):
    if compra_id:
        compra = get_object_or_404(Compras, id=compra_id)
        detalles = DetalleCompra.objects.filter(compra_id=compra_id)
    else:
        compra = None
        detalles = DetalleCompra.objects.all()
    return render(request, 'lista_detalles_compra.html', {'detalles': detalles, 'compra': compra})

def crear_detalle_compra(request, compra_id):
    compra = get_object_or_404(Compras, id=compra_id)
    if request.method == 'POST':
        form = DetalleCompraForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.compra = compra
            detalle.save()
            compra.actualizar_total()
            return redirect('crear_detalle_compra', compra_id=compra.id)
    else:
        form = DetalleCompraForm()
    detalles = DetalleCompra.objects.filter(compra=compra)
    return render(request, 'crear_detalle_compra.html', {'form': form, 'compra': compra, 'detalles': detalles})

def editar_detalle_compra(request, pk):
    detalle = get_object_or_404(DetalleCompra, pk=pk)
    if request.method == 'POST':
        form = DetalleCompraForm(request.POST, instance=detalle)
        if form.is_valid():
            detalle = form.save()
            return redirect('lista_detalles_compra', compra_id=detalle.compra.id)
    else:
        form = DetalleCompraForm(instance=detalle)
    return render(request, 'editar_detalle_compra.html', {'form': form, 'detalle': detalle})

def eliminar_detalle_compra(request, pk):
    detalle = get_object_or_404(DetalleCompra, pk=pk)
    if request.method == 'POST':
        compra_id = detalle.compra.id
        detalle.delete()
        return redirect('lista_detalles_compra', compra_id=compra_id)
    return render(request, 'eliminar_detalle_compra.html', {'detalle': detalle})

def exportar_excel(request):
    detalles_compra = DetalleCompra.objects.all()
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Detalles de Compra'

    headers = ['ID', 'Fecha', 'Proveedor', 'Producto', 'Cantidad', 'Precio Unitario', 'Total']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    for row_num, detalle in enumerate(detalles_compra, 2):
        ws[f'A{row_num}'] = detalle.id
        ws[f'B{row_num}'] = detalle.fecha.strftime('%d/%m/%Y')
        ws[f'C{row_num}'] = detalle.compra.proveedor.nombre if detalle.compra.proveedor else 'No especificado'
        ws[f'D{row_num}'] = detalle.producto
        ws[f'E{row_num}'] = detalle.cantidad
        ws[f'F{row_num}'] = detalle.precio_unitario
        ws[f'G{row_num}'] = detalle.total

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=detalles_compra.xlsx'
    wb.save(response)
    return response

