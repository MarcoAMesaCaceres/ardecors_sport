from django.shortcuts import render, get_object_or_404, redirect
from .models import Tarea
from .forms import TareaForm

# Exportar pdf, excel
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.utils import timezone
from django.contrib import messages

def lista_tareas(request):
    descripcion_query = request.GET.get('descripcion', '').strip()
    completada_query = request.GET.get('completada', '')
    fecha_creacion_query = request.GET.get('fecha_creacion', '').strip()
    fecha_vencimiento_query = request.GET.get('fecha_vencimiento', '').strip()

    tareas = Tarea.objects.all()

    if descripcion_query:
        tareas = tareas.filter(descripcion__icontains=descripcion_query)
    if completada_query:
        if completada_query.lower() == 'si':
            tareas = tareas.filter(completada=True)
        elif completada_query.lower() == 'no':
            tareas = tareas.filter(completada=False)
    if fecha_creacion_query:
        try:
            fecha_creacion = timezone.datetime.strptime(fecha_creacion_query, '%Y-%m-%d')
            tareas = tareas.filter(fecha_creacion__date=fecha_creacion)
        except ValueError:
            messages.error(request, "Formato de fecha de creación inválido. Use YYYY-MM-DD.")
    if fecha_vencimiento_query:
        try:
            fecha_vencimiento = timezone.datetime.strptime(fecha_vencimiento_query, '%Y-%m-%d')
            tareas = tareas.filter(fecha_vencimiento__date=fecha_vencimiento)
        except ValueError:
            messages.error(request, "Formato de fecha de vencimiento inválido. Use YYYY-MM-DD.")

    return render(request, 'lista_tareas.html', {'tareas': tareas})



def crear_tareas(request):
    if request.method == 'POST':
        form = TareaForm(request.POST)
        if form.is_valid():
            form.save()
            return redirect('lista_tareas')
    else:
        form = TareaForm()
    return render(request, 'crear_tareas.html', {'form': form})
    
def editar_tareas(request, pk):
    tareas = get_object_or_404(Tarea, pk=pk)
    if request.method == 'POST':
        form = TareaForm(request.POST, instance=tareas)
        if form.is_valid():
            form.save()
            return redirect('lista_tareas')
    else:
        form = TareaForm(instance=tareas)
    return render(request, 'editar_tareas.html', {'form': form, 'tareas': tareas})


def eliminar_tareas(request, pk):
    tareas = get_object_or_404(Tarea, pk=pk)
    if request.method == 'POST':
        tareas.delete()
        return redirect('lista_tareas')
    return render(request, 'eliminar_tareas.html', {'tareas': tareas})

def exportar_excel(request):
    tareas = Tarea.objects.all()
    
    # Crear libro y hoja de Excel
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'tareas'

    # Agregar encabezados
    headers = ['ID', 'Descripcion', 'FechaCreacion', 'FechaVencimiento', 'Completada', 'Acciones']
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        ws[f'{col_letter}1'] = header

    # Agregar filas de ventas
    for row_num, venta in enumerate(tareas, 2):
        ws[f'A{row_num}'] = tareas.id
        ws[f'F{row_num}'] = tareas.Descripcion
        ws[f'B{row_num}'] = tareas.FechaCreacion.strftime('%d/%m/%Y')
        ws[f'C{row_num}'] = tareas.FechaVencimiento.strftime('%d/%m/%Y')
        ws[f'E{row_num}'] = tareas.Completada
        ws[f'D{row_num}'] = tareas.Acciones.nombre if tareas.Acciones else 'No especificado'
        

    # Crear respuesta HTTP para descargar el archivo
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'
    wb.save(response)
    return response

def exportar_pdf(request):
    # Crear un buffer de bytes para el PDF
    buffer = BytesIO()

    # Crear el documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Obtener los datos de las ventas
    tareas = Tarea.objects.all()

    # Crear los datos para la tabla
    data = [['ID', 'Descripcion', 'FechaCreacion', 'FechaVencimiento', 'Completada', 'Acciones']]  # Encabezados
    for tareas in tareas:
        data.append([str(tareas.id), tareas.Descripcion, str(tareas.FechaCreacion), str(tareas.FechaVencimiento), str(tareas.Completada) , str(tareas.Acciones)])

    # Crear la tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Agregar la tabla al documento
    elements.append(table)

    # Construir el PDF
    doc.build(elements)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='tareas.pdf')

