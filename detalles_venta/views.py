from django.shortcuts import get_object_or_404, redirect, render
from .models import DetalleVenta, Venta

from django.contrib import messages
from io import BytesIO

# importacion para exportar archivos como pdf, exce
from django.http import HttpResponse
from io import BytesIO
import openpyxl
from openpyxl.utils import get_column_letter
from django.http import FileResponse
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle
from django.core.exceptions import ValidationError
from django.shortcuts import render
from django.db.models import Q
from .models import Venta, DetalleVenta
from .forms import  DetalleVentaSearchForm
from .forms import DetalleVentaForm
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import xlsxwriter


def lista_detalles_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    form = DetalleVentaSearchForm(request.GET)
    detalles = DetalleVenta.objects.filter(venta=venta)

    if form.is_valid():
        articulo = form.cleaned_data.get('articulo')
        cantidad_min = form.cleaned_data.get('cantidad_min')
        cantidad_max = form.cleaned_data.get('cantidad_max')
        precio_unitario_min = form.cleaned_data.get('precio_unitario_min')
        precio_unitario_max = form.cleaned_data.get('precio_unitario_max')

        if articulo:
            detalles = detalles.filter(articulo=articulo)
        if cantidad_min is not None:
            detalles = detalles.filter(cantidad__gte=cantidad_min)
        if cantidad_max is not None:
            detalles = detalles.filter(cantidad__lte=cantidad_max)
        if precio_unitario_min is not None:
            detalles = detalles.filter(precio_unitario__gte=precio_unitario_min)
        if precio_unitario_max is not None:
            detalles = detalles.filter(precio_unitario__lte=precio_unitario_max)
    
    return render(request, 'lista_detalles_venta.html', {'venta': venta, 'detalles': detalles, 'form': form})

def crear_detalle_venta(request, venta_id):
    venta = get_object_or_404(Venta, pk=venta_id)
    if request.method == 'POST':
        form = DetalleVentaForm(request.POST)
        if form.is_valid():
            detalle = form.save(commit=False)
            detalle.venta = venta
            detalle.save()
            messages.success(request, 'Detalle de venta agregado exitosamente.')
            return redirect('lista_detalles_venta', venta_id=venta.id)
    else:
        form = DetalleVentaForm()
    return render(request, 'crear_detalle_venta.html', {'form': form, 'venta': venta})

def editar_detalle_venta(request, pk):
    detalle = get_object_or_404(DetalleVenta, pk=pk)
    if request.method == 'POST':
        form = DetalleVentaForm(request.POST, instance=detalle)
        if form.is_valid():
            try:
                form.save()
                messages.success(request, 'Detalle de venta actualizado exitosamente.')
                return redirect('lista_detalles_venta', venta_id=detalle.venta.id)
            except ValidationError as e:
                for field, errors in e.message_dict.items():
                    for error in errors:
                        form.add_error(field, error)
    else:
        form = DetalleVentaForm(instance=detalle)
    return render(request, 'editar_detalle_venta.html', {'form': form, 'detalle': detalle})
def eliminar_detalle_venta(request, pk):
    detalle = get_object_or_404(DetalleVenta, pk=pk)
    venta_id = detalle.venta.id
    if request.method == 'POST':
        detalle.delete()
        messages.success(request, 'Detalle de venta eliminado exitosamente.')
        return redirect('lista_detalles_venta', venta_id=venta_id)
    return render(request, 'eliminar_detalle_venta.html', {'detalle': detalle})

def exportar_pdf(request):
    # Crear un buffer de bytes para el PDF
    buffer = BytesIO()

    # Crear el documento PDF
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []

    # Obtener los datos de las ventas
    detalleVenta = DetalleVenta.objects.all()

    # Crear los datos para la tabla
    data = [['ID', 'Cliente', 'Fecha', 'Total']]  # Encabezados
    for detalleVenta in DetalleVenta:
        data.append([str(detalleVenta.id), detalleVenta.cliente, str(detalleVenta.fecha), str(detalleVenta.total)])

    # Crear la tabla
    table = Table(data)
    
    # Estilo de la tabla
    style = TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 14),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 12),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ])
    table.setStyle(style)

    # Agregar la tabla al documento
    elements.append(table)

    # Construir el PDF
    doc.build(elements)

    # FileResponse sets the Content-Disposition header so that browsers
    # present the option to save the file.
    buffer.seek(0)
    return FileResponse(buffer, as_attachment=True, filename='ventas.pdf')

def exportar_excel(request):
    # Crear un buffer de bytes para el archivo Excel
    buffer = BytesIO()

    # Crear un nuevo libro de trabajo de Excel y agregar una hoja
    workbook = xlsxwriter.Workbook(buffer)
    worksheet = workbook.add_worksheet()

    # Definir estilos
    titulo_estilo = workbook.add_format({
        'bold': True,
        'font_size': 14,
        'align': 'center',
        'valign': 'vcenter',
        'bg_color': '#4F81BD',
        'font_color': 'white'
    })

    dato_estilo = workbook.add_format({
        'align': 'center',
        'valign': 'vcenter'
    })

    # Escribir los encabezados
    encabezados = ['ID', 'Cliente', 'Fecha', 'Total']
    for col, encabezado in enumerate(encabezados):
        worksheet.write(0, col, encabezado, titulo_estilo)

    # Obtener los datos de las ventas
    detalles_venta = DetalleVenta.objects.all()

    # Escribir los datos en el Excel
    for row, detalle in enumerate(detalles_venta, start=1):
        worksheet.write(row, 0, detalle.id, dato_estilo)
        worksheet.write(row, 1, detalle.cliente, dato_estilo)
        worksheet.write(row, 2, detalle.fecha.strftime('%Y-%m-%d'), dato_estilo)
        worksheet.write(row, 3, detalle.total, dato_estilo)

    # Ajustar el ancho de las columnas
    for i, _ in enumerate(encabezados):
        worksheet.set_column(i, i, 15)

    # Cerrar el libro de trabajo
    workbook.close()

    # Preparar la respuesta HTTP
    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=ventas.xlsx'

    return response
